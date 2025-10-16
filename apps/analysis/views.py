from django.http import HttpResponse
from django.views import View
import io
import os
from openpyxl import load_workbook
import pandas as pd
import chardet


class FileUploadView(View):
    def post(self, request):
        # Проверяем наличие файла
        if 'file_field' not in request.FILES:
            return HttpResponse("Файл не найден", status=400)

        uploaded_file = request.FILES['file_field']

        # Создаем бинарный поток
        binary_buffer = io.BytesIO()

        # Читаем и сохраняем файл в бинарный поток
        for chunk in uploaded_file.chunks():
            binary_buffer.write(chunk)

        binary_buffer.seek(0)

        # Определяем тип файла
        file_type = self.detect_file_type(uploaded_file.name, binary_buffer)

        if file_type == 'unknown':
            return HttpResponse("Неподдерживаемый формат файла. Используйте XLSX или CSV.", status=400)

        # Обрабатываем файл в зависимости от типа
        try:
            if file_type == 'xlsx':
                return self.process_excel(binary_buffer)
            elif file_type == 'csv':
                return self.process_csv(binary_buffer)
        except Exception as e:
            return HttpResponse(f"Ошибка обработки файла: {str(e)}", status=400)

    def detect_file_type(self, filename, binary_buffer):
        """
        Определяет тип файла по расширению и содержимому
        """
        # Определяем по расширению
        ext = os.path.splitext(filename)[1].lower()

        if ext == '.xlsx':
            return 'xlsx'
        elif ext == '.csv':
            return 'csv'

        # Если расширение неопределено, определяем по содержимому
        return self.detect_by_content(binary_buffer)

    def detect_by_content(self, binary_buffer):
        """
        Определяет тип файла по содержимому
        """
        # Сохраняем текущую позицию
        current_pos = binary_buffer.tell()
        binary_buffer.seek(0)

        try:
            # Читаем первые байты для анализа
            first_bytes = binary_buffer.read(8)
            binary_buffer.seek(current_pos)  # Возвращаем позицию

            # XLSX файлы начинаются с PK (ZIP архив)
            if first_bytes.startswith(b'PK\x03\x04'):
                return 'xlsx'

            # Пробуем определить кодировку CSV
            binary_buffer.seek(0)
            sample = binary_buffer.read(1024)
            binary_buffer.seek(current_pos)

            # Определяем кодировку
            encoding_info = chardet.detect(sample)
            encoding = encoding_info['encoding']

            if encoding and self.is_likely_csv(sample, encoding):
                return 'csv'

        except Exception:
            pass

        return 'unknown'

    def is_likely_csv(self, sample_bytes, encoding):
        """
        Проверяет, похожи ли данные на CSV
        """
        try:
            sample_text = sample_bytes.decode(encoding, errors='ignore')

            # Проверяем наличие типичных CSV разделителей
            lines = sample_text.split('\n')
            if lines:
                first_line = lines[0].strip()
                # Считаем количество запятых, точек с запятой или табов
                comma_count = first_line.count(',')
                semicolon_count = first_line.count(';')
                tab_count = first_line.count('\t')

                # Если есть несколько разделителей - вероятно CSV
                if comma_count > 1 or semicolon_count > 1 or tab_count > 1:
                    return True

        except Exception:
            pass

        return False

def process_excel(self, binary_buffer):
    """
    Обрабатывает XLSX файл
    """
    try:
        # Используем openpyxl для работы с Excel
        workbook = load_workbook(binary_buffer)
        sheet = workbook.active

        # Пример: читаем данные из первого листа
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        return HttpResponse(f"XLSX файл обработан. Строк: {len(data)}")

    except Exception as e:
        raise Exception(f"Ошибка чтения XLSX: {str(e)}")


def process_csv(self, binary_buffer):
    """
    Обрабатывает CSV файл
    """
    try:
        # Определяем кодировку
        binary_buffer.seek(0)
        sample = binary_buffer.read(1024)
        encoding_info = chardet.detect(sample)
        encoding = encoding_info['encoding'] or 'utf-8'

        binary_buffer.seek(0)

        # Используем pandas для чтения CSV
        df = pd.read_csv(binary_buffer, encoding=encoding)

        return HttpResponse(f"CSV файл обработан. Строк: {len(df)}, Колонок: {len(df.columns)}")

    except Exception as e:
        raise Exception(f"Ошибка чтения CSV: {str(e)}")